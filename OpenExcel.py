import openpyxl
filename = "F:\\Software testing_ETL_BI_Testing\\COUNT and NULL COUNT.xlsx"
workbook = openpyxl.load_workbook(filename)

sheet = workbook["Test Scenarios"]

# print(sheet.cell(row=2, column=2).value)
#
# for n in range(1, 17):
#     for a in range(1, 3):
#         print(sheet.cell(row=n, column=a).value)

sheet.cell(row=17, column=2).value = "Rohit Patil"
workbook.save(filename)
